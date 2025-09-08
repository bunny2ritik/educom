import re
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference


def extract_summary_tables_from_stream(stream):
    access_logs = []
    fticks_logs = []

    access_pattern = re.compile(
        r'(?P<timestamp>\w{3} \w{3} \d+ \d{2}:\d{2}:\d{2} \d{4}): '
        r'(?P<event>Access-(Accept|Reject)) for user (?P<user>[\w@.]+)? ?stationid (?P<stationid>[0-9a-fA-F:-]+).*?'
        r'from (?P<source>\S+)(?: to (?P<dest>\S+))?.*?\((?P<ip>[\d.]+)\)?(?:.*?cui (?P<cui>[a-f0-9]+))?',
        re.IGNORECASE)

    fticks_pattern = re.compile(
        r'(?P<timestamp>\w{3} \w{3} \d+ \d{2}:\d{2}:\d{2} \d{4}): '
        r'F-TICKS/eduroam/1.0#REALM=[^#]*#VISCOUNTRY=(?P<country>[^#]*)#VISINST=(?P<inst>[^#]*)#CSI=(?P<csi>[^#]*)#RESULT=(?P<result>[^#]*)#'
    )

    for line in stream:
        try:
            line = line.decode("utf-8", errors="ignore")
        except Exception:
            continue

        if (m := access_pattern.search(line)):
            access_logs.append({
                "Timestamp": m.group("timestamp"),
                "Event": m.group("event"),
                "Username": m.group("user"),
                "StationID": m.group("stationid"),
                "Source": m.group("source"),
                "Destination": m.group("dest"),
                "ServerIP": m.group("ip"),
                "CUI": m.group("cui")
            })

        if (m := fticks_pattern.search(line)):
            result = m.group("result")
            reason = "Authentication successful" if result == "OK" else "Authentication failed"
            fticks_logs.append({
                "Timestamp": m.group("timestamp"),
                "VISCOUNTRY": m.group("country"),
                "VISINST": m.group("inst"),
                "CSI": m.group("csi"),
                "RESULT": result,
                "Reason": reason
            })

    df_access = pd.DataFrame(access_logs)
    df_fticks = pd.DataFrame(fticks_logs)

    df_access["Timestamp"] = pd.to_datetime(df_access["Timestamp"], errors="coerce")
    df_fticks["Timestamp"] = pd.to_datetime(df_fticks["Timestamp"], errors="coerce")

    return df_access, df_fticks


def parse_and_generate_excel(log_data: str) -> BytesIO:
    df_access, df_fticks = extract_summary_tables_from_stream(BytesIO(log_data.encode("utf-8")))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_access.to_excel(writer, sheet_name="AccessLogs", index=False)
        df_fticks.to_excel(writer, sheet_name="FTicksLogs", index=False)

    output.seek(0)
    wb = load_workbook(output)

    ws_fticks = wb["FTicksLogs"]
    pie = PieChart()
    pie.title = "Auth Result (F-TICKS)"
    labels = Reference(ws_fticks, min_col=5, min_row=2, max_row=ws_fticks.max_row)
    data = Reference(ws_fticks, min_col=5, min_row=1, max_row=ws_fticks.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws_fticks.add_chart(pie, "H2")

    df_visinst = df_fticks["VISINST"].value_counts().reset_index()
    chart_ws = wb.create_sheet("ChartData")
    chart_ws.append(["VISINST", "Count"])
    for row in df_visinst.values:
        chart_ws.append(list(row))

    bar = BarChart()
    bar.title = "Top VISINSTs"
    data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=len(df_visinst)+1)
    label_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=len(df_visinst)+1)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(label_ref)
    ws_fticks.add_chart(bar, "H20")

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output